import pymysql
import json
from openpyxl import load_workbook

def SelSqlFun():
    """" This function will fetch the information from table for certain critiria() """
    with open(r"mysql_querry", "r") as file:
        data = json.load(file)
    db = pymysql.connect(host=data["DB_HOST"], user=data["DB_USER"], password=data["DB_PASSWD"], database=data["DB_NAME"], port=data["DB_PORT"])
    cursor = db.cursor()
    sql = f'SELECT {data["DB_FIELD"]} FROM {data["TABLE_NAME"]} WHERE app_code IN ({data["APP_CODE"]});'
    try:
       # Execute the SQL command
       cursor.execute(sql)
       # Fetch all the rows in a list of lists.
       results = cursor.fetchall()
       for row in results:
           if ( row[6] == 'PROD' or row[6] == 'DR'):
               env="prod"
               if( row[3] == "Windows" or row[3] == "Linux"):
                   infra="OS"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'F{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
               elif( row[3] == "Oracle" or row[3] == "SQLServer"):
                   infra="DB"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'G{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
           elif ( row[6] == 'DEV' or row[6] == 'QA'):
               env="nonprod"
               if( row[3] == "Windows" or row[3] == "Linux"):
                   infra="OS"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'D{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
               elif( row[3] == "Oracle" or row[3] == "SQLServer"):
                   infra="DB"
                   template1 = load_workbook(filename=f"template1_{env}_{infra}.xlsx")
                   main_sheet = template1.active
                   current_row = main_sheet.max_row
                   next_row = current_row + 1
                   main_sheet[f'B{next_row}'] = row[5]
                   main_sheet[f'C{next_row}'] = row[4]
                   template1.save(f"template1_{env}_{infra}.xlsx")
    except Exception as e:
       print (f"Error: unable to fetch data with error {e}")

    # disconnect from server
    db.close()
